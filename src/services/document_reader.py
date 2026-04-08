"""
Document reading service - Extract text from various file formats.

Provides utilities to read and extract text from PDF, DOCX, and XLSX files.
Includes file validation and size checking capabilities.
"""

import os
from typing import Optional
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook


class DocumentReader:
    """
    Service for reading and extracting text from different document types.
    
    Supports PDF, DOCX, and XLSX files with automatic format detection
    and comprehensive text extraction capabilities.
    """

    MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB in bytes

    @staticmethod
    def is_supported_file(file_path: str) -> bool:
        """
        Check if the file type is supported.
        
        Args:
            file_path: Path to the file
            
        Returns:
            True if file type is supported, False otherwise
        """
        _, ext = os.path.splitext(file_path.lower())
        return ext in ['.pdf', '.docx', '.xlsx']

    @staticmethod
    def check_file_size(file_path: str) -> bool:
        """
        Check if file size is within limits.
        
        Args:
            file_path: Path to the file
            
        Returns:
            True if file size is acceptable, False otherwise
        """
        return os.path.getsize(file_path) <= DocumentReader.MAX_FILE_SIZE

    @staticmethod
    def extract_text(file_path: str) -> str:
        """
        Extract text from the document based on its type.
        
        Args:
            file_path: Path to the document file
            
        Returns:
            Extracted text content
            
        Raises:
            ValueError: If file type is not supported
        """
        _, ext = os.path.splitext(file_path.lower())

        if ext == '.pdf':
            return DocumentReader._extract_pdf_text(file_path)
        elif ext == '.docx':
            return DocumentReader._extract_docx_text(file_path)
        elif ext == '.xlsx':
            return DocumentReader._extract_xlsx_text(file_path)
        else:
            raise ValueError("Unsupported file type")

    @staticmethod
    def _extract_pdf_text(file_path: str) -> str:
        """
        Extract text from PDF file.
        
        Args:
            file_path: Path to PDF file
            
        Returns:
            Extracted text from all pages
        """
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()

    @staticmethod
    def _extract_docx_text(file_path: str) -> str:
        """
        Extract text from DOCX file.
        
        Args:
            file_path: Path to DOCX file
            
        Returns:
            Extracted text from all paragraphs
        """
        doc = Document(file_path)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text.strip()

    @staticmethod
    def _extract_xlsx_text(file_path: str) -> str:
        """
        Extract text from XLSX file by concatenating all cell values.
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            Extracted text from all cells
        """
        wb = load_workbook(file_path)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + " "
                text += "\n"
        return text.strip()
